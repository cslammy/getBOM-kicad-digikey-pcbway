#!/usr/bin/env python3
"""
DigiKey BOM Filler — PCBWay Template
-------------------------------------
Reads a KiCad/EDA CSV export, searches DigiKey for part numbers,
and outputs a filled PCBWay BOM Excel file.

Usage:
    python digikey_bom.py input.csv [output.xlsx]

Requirements:
    pip install pandas openpyxl requests beautifulsoup4

DigiKey API (optional but recommended):
    Set environment variables DIGIKEY_CLIENT_ID and DIGIKEY_CLIENT_SECRET,
    or pass --client-id / --client-secret on the command line.
    Get free credentials at https://developer.digikey.com
"""

import argparse
import os
import re
import shutil
import sys
import time
import urllib.parse
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

TEMPLATE_PATH = Path(__file__).parent / "Sample_BOM_PCBWay.xlsx"

DATA_START_ROW = 7   # Row where component data begins in the PCBWay template
COL_ITEM       = 1   # A — Item #
COL_DESIGNATOR = 2   # B — *Designator
COL_QTY        = 3   # C — *Qty
COL_MFR        = 4   # D — Manufacturer
COL_MFG_PN     = 5   # E — *Mfg Part #
COL_DESC       = 6   # F — Description / Value
COL_PACKAGE    = 7   # G — *Package/Footprint
COL_TYPE       = 8   # H — Type
COL_NOTES      = 9   # I — Your Instructions / Notes  ← DigiKey PN or supplier URL goes here

# Parts that are auto-filled with a preferred supplier URL — no prompting needed
PREFERRED_PARTS = [
    {
        "keywords": ["europower", "euro power", "eurorack power"],
        "url": "https://www.taydaelectronics.com/10-pin-box-header-connector-2-54mm.html",
        "description": "Eurorack 10-pin power connector",
    },
]

# Parts that need a manual URL from the user (checked against Reference + Value + Footprint)
MANUAL_PART_KEYWORDS = [
    "potentiometer", "pot", "trimmer", "rheostat",
    "switch", " sw ", "pushbutton", "toggle", "tact",
    "knob",
    "connector", "conn", "header", "socket", "terminal block",
    "jack",
    "europower", "euro power", "eurorack power",
]

# Parts that should be silently skipped — not real purchasable components
SKIP_KEYWORDS = [
    "mounthole", "mount_hole", "mount hole", "testpoint", "test_point",
    "fiducial", "dnp", "no_connect", "noconnect",
]

def should_skip(ref: str, value: str, package: str) -> bool:
    """Return True if this row is a non-purchasable item like a mount hole."""
    text = " ".join([ref, value, package]).lower().replace("-", "_")
    return any(kw in text for kw in SKIP_KEYWORDS)

# Column name aliases for mapping CSV columns → PCBWay fields
COLUMN_ALIASES = {
    "designator":    ["reference", "designator", "ref", "refdes"],
    "qty":           ["qty", "quantity", "count"],
    "manufacturer":  ["manufacturer", "mfr", "mfg"],
    "mfg_pn":        ["mfg part #", "mfg part", "mpn", "manufacturer part", "part number", "partnumber"],
    "description":   ["description", "value", "desc", "component"],
    "package":       ["package", "footprint", "case"],
    "type":          ["type", "mount"],
}

SCRAPE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------

def map_columns(df: pd.DataFrame) -> dict:
    """Return a mapping of PCBWay field → actual CSV column name."""
    cols_lower = {c.lower().strip(): c for c in df.columns}
    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in cols_lower:
                mapping[field] = cols_lower[alias]
                break
    return mapping


def get_field(row, mapping: dict, field: str):
    col = mapping.get(field)
    if col is None:
        return None
    val = row.get(col)
    if pd.isna(val) or str(val).strip() in ("~", ""):
        return None
    return str(val).strip()


# ---------------------------------------------------------------------------
# Part classification
# ---------------------------------------------------------------------------

def search_text(row, mapping: dict) -> str:
    """Combine reference + value + footprint into one searchable string."""
    parts = [
        get_field(row, mapping, "designator") or "",
        get_field(row, mapping, "description") or "",
        get_field(row, mapping, "package") or "",
    ]
    return " ".join(parts).lower()


def preferred_url(text: str) -> str | None:
    """Return a preferred supplier URL if this part matches a known entry."""
    for entry in PREFERRED_PARTS:
        if any(kw in text for kw in entry["keywords"]):
            return entry["url"]
    return None


def needs_manual_url(text: str) -> bool:
    """Return True if this part should be sourced manually."""
    return any(kw in text for kw in MANUAL_PART_KEYWORDS)


# ---------------------------------------------------------------------------
# DigiKey API search
# ---------------------------------------------------------------------------

def get_digikey_token(client_id: str, client_secret: str) -> str:
    resp = requests.post(
        "https://api.digikey.com/v1/oauth2/token",
        data={
            "client_id": client_id,
            "client_secret": client_secret,
            "grant_type": "client_credentials",
        },
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def search_digikey_api(query: str, client_id: str, token: str) -> str | None:
    headers = {
        "Authorization": f"Bearer {token}",
        "X-DIGIKEY-Client-Id": client_id,
        "X-DIGIKEY-Locale-Site": "US",
        "X-DIGIKEY-Locale-Language": "en",
        "X-DIGIKEY-Locale-Currency": "USD",
        "Content-Type": "application/json",
    }
    payload = {
        "keywords": query,
        "limit": 1,
        "offset": 0,
        "filterOptionsRequest": {"inStock": True},
    }
    try:
        resp = requests.post(
            "https://api.digikey.com/products/v4/search/keyword",
            json=payload,
            headers=headers,
            timeout=15,
        )
        if resp.status_code == 429:
            print("  [rate limited] waiting 2s...")
            time.sleep(2)
            resp = requests.post(
                "https://api.digikey.com/products/v4/search/keyword",
                json=payload,
                headers=headers,
                timeout=15,
            )
        if resp.status_code != 200:
            print(f"  [API error] HTTP {resp.status_code}: {resp.text[:200]}")
            return None

        data = resp.json()

        # DEBUG: set DIGIKEY_DEBUG=1 env var to print raw API response
        if os.environ.get("DIGIKEY_DEBUG"):
            import json as _json
            print(f"  [DEBUG] Raw response: {_json.dumps(data, indent=2)[:1000]}")

        # Try multiple response shapes — DigiKey has varied this across API versions
        products = (
            data.get("Products") or
            data.get("products") or
            data.get("ProductsV4") or
            []
        )

        if not products:
            products = (
                data.get("exactManufacturerProducts") or
                data.get("nonExactOrderableProducts") or
                []
            )

        if not products:
            print(f"  [API] No products found. Response keys: {list(data.keys())}")
            return None

        product = products[0]

        # Try all known locations for the DigiKey part number
        pn = (
            product.get("DigiKeyPartNumber") or
            product.get("digiKeyPartNumber") or
            product.get("PartNumber") or
            product.get("partNumber") or
            None
        )

        # v4 API sometimes nests it under productVariations
        if not pn:
            variations = (
                product.get("productVariations") or
                product.get("ProductVariations") or
                []
            )
            if variations:
                pn = (
                    variations[0].get("digiKeyProductNumber") or
                    variations[0].get("DigiKeyProductNumber") or
                    variations[0].get("partNumber") or
                    None
                )

        if not pn:
            print(f"  [API] Found product but no PN extracted. Product keys: {list(product.keys())}")

        # Extract manufacturer name and part number from response
        mfr_name = None
        mfr_pn = None
        mfr_block = product.get("Manufacturer") or product.get("manufacturer")
        if isinstance(mfr_block, dict):
            mfr_name = mfr_block.get("Name") or mfr_block.get("name")
        elif isinstance(mfr_block, str):
            mfr_name = mfr_block
        mfr_pn = (
            product.get("ManufacturerProductNumber") or
            product.get("manufacturerProductNumber") or
            product.get("MfgPartNumber") or
            None
        )

        return {"digikey_pn": pn, "manufacturer": mfr_name, "mfg_pn": mfr_pn}

    except Exception as e:
        print(f"  [API error] {e}")
    return None


# ---------------------------------------------------------------------------
# DigiKey web scrape fallback
# ---------------------------------------------------------------------------

def search_digikey_scrape(query: str) -> str | None:
    q = urllib.parse.quote_plus(query)
    url = f"https://www.digikey.com/en/products/result?keywords={q}"
    time.sleep(0.6)
    try:
        resp = requests.get(url, headers=SCRAPE_HEADERS, timeout=12)
        if resp.status_code == 403 or "captcha" in resp.text.lower():
            print("  [scrape blocked] DigiKey returned 403/captcha")
            return None
        soup = BeautifulSoup(resp.text, "html.parser")
        links = soup.select("a[href*='/en/products/detail/']")
        if links:
            href = links[0]["href"].split("?")[0].rstrip("/")
            pn = href.split("/")[-1]
            if pn:
                return {"digikey_pn": pn, "manufacturer": None, "mfg_pn": None}
    except Exception as e:
        print(f"  [scrape error] {e}")
    return None


# ---------------------------------------------------------------------------
# Main search dispatcher
# ---------------------------------------------------------------------------

def search_digikey(query: str, client_id: str | None, token: str | None) -> dict | None:
    """Returns dict with keys: digikey_pn, manufacturer, mfg_pn — or None if not found."""
    if client_id and token:
        return search_digikey_api(query, client_id, token)
    return search_digikey_scrape(query)


def best_search_query(mfg_pn: str | None, description: str | None, package: str | None = None) -> str:
    """Build the best search query. Include package when description is a bare value like 10k, .1uF."""
    if mfg_pn and mfg_pn.strip():
        return mfg_pn.strip()
    desc = (description or "").strip()
    if not desc:
        return ""
    # If description looks like a bare value (short, no spaces, or just a net name),
    # append the package to help DigiKey narrow results
    looks_generic = len(desc) <= 8 or " " not in desc
    if looks_generic and package and package.strip():
        # Extract just the package identifier, not the full KiCad footprint path
        pkg = package.split(":")[-1].split("_")[0]  # e.g. "0603" from "Resistor_SMD:R_0603_..."
        return f"{desc} {pkg}".strip()
    return desc


# ---------------------------------------------------------------------------
# Write PCBWay template
# ---------------------------------------------------------------------------

def write_bom(components: list[dict], output_path: Path):
    if not TEMPLATE_PATH.exists():
        sys.exit(
            f"ERROR: PCBWay template not found at {TEMPLATE_PATH}\n"
            "Place Sample_BOM_PCBWay.xlsx in the same folder as this script."
        )
    shutil.copy(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)
    ws = wb["Sheet1"]

    # Unmerge any merged cells in the data area
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= DATA_START_ROW:
            ws.unmerge_cells(str(rng))

    # Clear existing sample rows
    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col in range(1, 10):
            ws.cell(row=row, column=col).value = None

    for i, c in enumerate(components):
        row = DATA_START_ROW + i
        ws.cell(row=row, column=COL_ITEM).value       = i + 1
        ws.cell(row=row, column=COL_DESIGNATOR).value = c.get("designator")
        ws.cell(row=row, column=COL_QTY).value        = c.get("qty")
        ws.cell(row=row, column=COL_MFR).value        = c.get("manufacturer")
        ws.cell(row=row, column=COL_MFG_PN).value     = c.get("mfg_pn")
        ws.cell(row=row, column=COL_DESC).value       = c.get("description")
        ws.cell(row=row, column=COL_PACKAGE).value    = c.get("package")
        ws.cell(row=row, column=COL_TYPE).value       = c.get("type", "SMD")
        ws.cell(row=row, column=COL_NOTES).value      = c.get("notes")

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Fill a PCBWay BOM from a KiCad CSV export.")
    parser.add_argument("input_csv", help="Path to input CSV file")
    parser.add_argument("output_xlsx", nargs="?", default="BOM_filled.xlsx",
                        help="Output Excel file (default: BOM_filled.xlsx)")
    parser.add_argument("--client-id",     default=os.environ.get("DIGIKEY_CLIENT_ID"),
                        help="DigiKey API Client ID (or set DIGIKEY_CLIENT_ID env var)")
    parser.add_argument("--client-secret", default=os.environ.get("DIGIKEY_CLIENT_SECRET"),
                        help="DigiKey API Client Secret (or set DIGIKEY_CLIENT_SECRET env var)")
    args = parser.parse_args()

    # --- Load CSV ---
    print(f"\nReading {args.input_csv}...")
    df = pd.read_csv(args.input_csv)
    mapping = map_columns(df)
    print(f"Columns detected: {dict(mapping)}")

    if "designator" not in mapping and "description" not in mapping:
        sys.exit("ERROR: Could not identify a Designator or Description column. "
                 "Please rename your CSV columns and try again.")

    # --- DigiKey auth ---
    token = None
    client_id = args.client_id
    if client_id and args.client_secret:
        print("\nAuthenticating with DigiKey API...")
        try:
            token = get_digikey_token(client_id, args.client_secret)
            print("  DigiKey API token obtained.")
        except Exception as e:
            print(f"  WARNING: API auth failed ({e}). Falling back to web scraping.")
            client_id = None
    else:
        print("\nNo DigiKey API credentials found — using web scraping fallback.")
        print("(For better results: set DIGIKEY_CLIENT_ID and DIGIKEY_CLIENT_SECRET env vars)")

    # --- Classify parts ---
    auto_rows    = []  # Will be searched on DigiKey
    manual_rows  = []  # Need a URL from the user
    preferred_rows = []  # Known URL, no prompting needed

    for _, row in df.iterrows():
        text = search_text(row, mapping)
        pref_url = preferred_url(text)
        if pref_url:
            preferred_rows.append((row, pref_url))
        elif needs_manual_url(text):
            manual_rows.append(row)
        else:
            auto_rows.append(row)

    print(f"\n{len(auto_rows)} parts will be searched on DigiKey automatically.")
    print(f"{len(preferred_rows)} parts have a known preferred supplier URL.")
    print(f"{len(manual_rows)} parts need a supplier URL from you.")

    # --- Collect manual URLs ---
    manual_urls = {}
    if manual_rows:
        print("\n" + "="*60)
        print("The following parts are typically better sourced outside")
        print("DigiKey (e.g. Tayda, Thonk, Mouser).")
        print("Please provide a product URL for each, or press Enter to skip.")
        print("="*60)
        for i, row in enumerate(manual_rows):
            ref  = get_field(row, mapping, "designator") or "?"
            desc = get_field(row, mapping, "description") or "?"
            pkg  = get_field(row, mapping, "package") or ""
            print(f"\n{i+1}. {ref} — {desc} ({pkg})")
            while True:
                url = input("   URL (or Enter to skip): ").strip()
                if not url:
                    manual_urls[ref] = None
                    break
                if url.startswith("http://") or url.startswith("https://"):
                    manual_urls[ref] = url
                    break
                print("   Please enter a valid URL starting with http:// or https://")

    # --- Search DigiKey for auto parts ---
    print("\n" + "="*60)
    print("Searching DigiKey for standard parts...")
    print("="*60)

    auto_results = {}
    not_found = []

    for row in auto_rows:
        ref  = get_field(row, mapping, "designator") or "?"
        mpn  = get_field(row, mapping, "mfg_pn")
        desc = get_field(row, mapping, "description")
        pkg  = get_field(row, mapping, "package")
        # Skip non-purchasable items
        if should_skip(ref, desc or "", pkg or ""):
            print(f"  Skipping: {ref} — {desc} (mount hole / non-purchasable)")
            continue
        query = best_search_query(mpn, desc, pkg)
        if not query:
            print(f"  Skipping: {ref} — no searchable value")
            continue
        print(f"  Searching: {ref} — {query}")
        result = search_digikey(query, client_id, token)
        if result and result.get("digikey_pn"):
            print(f"    → {result['digikey_pn']} | {result.get('manufacturer','?')} | {result.get('mfg_pn','?')}")
            auto_results[ref] = result
        else:
            print(f"    → not found")
            not_found.append((ref, desc))
        time.sleep(0.3)

    # --- Assemble components list ---
    components = []

    def make_component(row, notes, api_result=None):
        # Use manufacturer/mfg_pn from DigiKey API if available, fall back to CSV values
        mfr = (api_result or {}).get("manufacturer") or get_field(row, mapping, "manufacturer")
        mfg_pn = (api_result or {}).get("mfg_pn") or get_field(row, mapping, "mfg_pn")
        return {
            "designator":   get_field(row, mapping, "designator"),
            "qty":          get_field(row, mapping, "qty"),
            "manufacturer": mfr,
            "mfg_pn":       mfg_pn,
            "description":  get_field(row, mapping, "description"),
            "package":      get_field(row, mapping, "package"),
            "type":         get_field(row, mapping, "type") or "SMD",
            "notes":        notes,
        }

    # Preserve original CSV row order
    for _, row in df.iterrows():
        text = search_text(row, mapping)
        ref  = get_field(row, mapping, "designator") or "?"
        pref_url = preferred_url(text)
        if pref_url:
            components.append(make_component(row, pref_url))
        elif needs_manual_url(text):
            components.append(make_component(row, manual_urls.get(ref)))
        else:
            api_result = auto_results.get(ref)
            notes = api_result.get("digikey_pn") if isinstance(api_result, dict) else None
            components.append(make_component(row, notes, api_result=api_result))

    # --- Write output ---
    output_path = Path(args.output_xlsx)
    print(f"\nWriting {output_path}...")
    write_bom(components, output_path)

    # --- Summary ---
    filled   = sum(1 for c in components if c["notes"])
    total    = len(components)
    print(f"\n{'='*60}")
    print(f"Done! {filled}/{total} parts filled.")
    print(f"Output: {output_path.resolve()}")
    if not_found:
        print(f"\n{len(not_found)} part(s) not found on DigiKey — left blank:")
        for ref, desc in not_found:
            print(f"  - {ref}: {desc}")
    print("="*60)


if __name__ == "__main__":
    main()
